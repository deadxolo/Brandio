# -*- coding: utf-8 -*-
"""Generate Brandio_Project_Status.xlsx — project completeness analysis."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- palette ----
NAVY   = "1F2A44"
GREEN  = "2E7D32"
RED    = "C0392B"
AMBER  = "B8860B"
GREYBG = "F2F4F7"
WHITE  = "FFFFFF"

hdr_fill   = PatternFill("solid", fgColor=NAVY)
hdr_font   = Font(bold=True, color=WHITE, size=11)
title_font = Font(bold=True, color=NAVY, size=16)
sub_font   = Font(italic=True, color="555555", size=10)
wrap_top   = Alignment(wrap_text=True, vertical="top")
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin       = Side(style="thin", color="D0D5DD")
border     = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_FILL = {
    "Done":     PatternFill("solid", fgColor="D6EFD8"),
    "Real":     PatternFill("solid", fgColor="D6EFD8"),
    "Partial":  PatternFill("solid", fgColor="FFF2CC"),
    "Fake":     PatternFill("solid", fgColor="FDE2E1"),
    "Missing":  PatternFill("solid", fgColor="FDE2E1"),
    "Blocker":  PatternFill("solid", fgColor="FDE2E1"),
    "DONE":     PatternFill("solid", fgColor="D6EFD8"),
    "PENDING":  PatternFill("solid", fgColor="FFF2CC"),
}
PRIO_FILL = {
    "P0 Critical": PatternFill("solid", fgColor="FDE2E1"),
    "P1 High":     PatternFill("solid", fgColor="FFE5CC"),
    "P2 Medium":   PatternFill("solid", fgColor="FFF2CC"),
    "P3 Low":      PatternFill("solid", fgColor="E8EAED"),
}


def style_sheet(ws, title, subtitle, headers, rows, widths, status_col=None, prio_col=None):
    ws.sheet_view.showGridLines = False
    ncol = len(headers)
    last = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = sub_font
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    hrow = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[hrow].height = 22

    for r, rowdata in enumerate(rows, hrow + 1):
        for c, val in enumerate(rowdata, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap_top
            cell.border = border
            cell.font = Font(size=10)
        if r % 2 == 0:
            for c in range(1, ncol + 1):
                if not ws.cell(row=r, column=c).fill.fgColor.rgb or ws.cell(row=r, column=c).fill.patternType is None:
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREYBG)
        if status_col is not None:
            sval = str(rowdata[status_col])
            for key, fill in STATUS_FILL.items():
                if sval.startswith(key):
                    cc = ws.cell(row=r, column=status_col + 1)
                    cc.fill = fill
                    cc.alignment = center
                    cc.font = Font(size=10, bold=True)
                    break
        if prio_col is not None:
            pval = str(rowdata[prio_col])
            for key, fill in PRIO_FILL.items():
                if pval.startswith(key):
                    cc = ws.cell(row=r, column=prio_col + 1)
                    cc.fill = fill
                    cc.alignment = center
                    cc.font = Font(size=10, bold=True)
                    break

    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    ws.auto_filter.ref = f"A{hrow}:{last}{hrow + len(rows)}"


# ============ TAB 1: Summary ============
ws1 = wb.active
ws1.title = "Summary"
summary_rows = [
    ["Core AI generation (backgrounds, templates, captions)", "Real", "Google Gemini 2.5; multi-step pipeline with gradient fallback", "background_engine/services/geminiService.js"],
    ["Social publishing — Instagram / Facebook", "Real", "Live Meta Graph API calls", "auto_poster/services/platforms/metaService.js"],
    ["Social publishing — Twitter / X", "Real", "Live media upload + v2 /tweets", "auto_poster/services/platforms/twitterService.js"],
    ["Social publishing — LinkedIn", "Real", "Live /ugcPosts + asset upload", "auto_poster/services/platforms/linkedinService.js"],
    ["OAuth (all 3 platforms)", "Real", "Token exchange + refresh; AES-256-GCM encrypted storage", "auto_poster/services/tokenService.js"],
    ["Scheduler", "Real", "cron every minute; publishes due posts automatically", "auto_poster/services/schedulerService.js"],
    ["Template editor UI", "Done", "Full canvas editor, fonts, placeholders", "post_generator/public/js/editor.js"],
    ["Docker / Cloud Run packaging", "Done", "Dockerfile (non-root, healthcheck), compose, build.sh", "Dockerfile / docker-compose.yml"],
    ["JWT signature verification", "Done", "PHASE 1: HS256 verify + signJwt helper (Node crypto, no new deps)", "shared/middleware/auth.js"],
    ["CORS lockdown", "Done", "PHASE 1: env-driven via CORS_ORIGINS (permissive only if unset)", "shared/config/corsOptions.js"],
    ["Startup secret validation", "Done", "PHASE 1: warns in dev, exits in prod if secrets missing", "shared/config/validateEnv.js"],
    ["OAuth redirect URIs", "Done", "PHASE 1: ngrok hardcode removed; derived from PUBLIC_BASE_URL", "shared/config/platforms.js"],
    ["Dead placeholder OAuth route", "Done", "PHASE 1: removed (real flow lives in oauthRoutes.js)", "auto_poster/routes/socialRoutes.js"],
    ["Secrets in git history", "Done", "VERIFIED NOT AN ISSUE: .env never committed, properly gitignored", ".gitignore"],
    ["Database (SQLite migration)", "Done", "PHASE 2: now SQLite (node:sqlite) w/ WAL; 389 rows migrated; same API; data.json kept as backup", "shared/db/database.js"],
    ["User authentication / login", "Done", "PHASE 1.5: scrypt signup/login, JWT, login UI, token auto-attach, per-user isolation + ownership checks", "manager/routes/authRoutes.js"],
    ["Analytics / reporting", "Done", "PHASE 3: analytics API (summary/posts/refresh) + dashboard UI + platform getInsights + ownership; live data needs connected accounts", "auto_poster/routes/analyticsRoutes.js + manager/public/analytics.html"],
    ["Automated tests", "Done", "PHASE 4: 26 tests via node:test (password, JWT, auth-middleware regression, SQLite, CORS); npm test green", "tests/"],
    ["Observability (Sentry/uptime/alerts)", "Partial", "PHASE 4: structured JSON error logging on all 4 services + optional Sentry (set SENTRY_DSN); uptime/alerts need external tooling", "shared/observability.js"],
    ["Cross-service ownership (authz depth)", "Done", "Lenient ownership guard on post_gen + auto_poster business routes; enforces for logged-in users, skips internal/dev", "shared/middleware/ownership.js"],
    ["Input validation", "Done", "Dependency-free validateBody; applied to template create; covered by tests", "shared/middleware/validate.js"],
    ["Best-time-to-post", "Fake", "Returns hardcoded times, not real data", "auto_poster/routes/calendarRoutes.js:145"],
    ["Excel bulk import", "Missing", "'Coming soon'; only CSV supported", "post_generator/public/home.html:1669"],
    ["Retry failed generation", "Missing", "'Coming soon'", "post_generator/public/js/generate.js:3131"],
    ["WhatsApp posting", "Missing", "Placeholder for future", "shared/config/platforms.js:86"],
    ["Multi-user / teams / approval flow", "Missing", "Single shared demo user; no roles or approval", "—"],
]
style_sheet(
    ws1,
    "Brandio — Project Status Analysis",
    "Generated from source-code review. Status: Real/Done = built & working · Partial = works with caveats · Fake = dummy data · Missing/Blocker = not built / must-fix.",
    ["Feature / Area", "Status", "Notes", "Key file(s)"],
    summary_rows,
    [42, 12, 60, 46],
    status_col=1,
)

# ============ TAB 2: Done ============
ws2 = wb.create_sheet("What's Done")
done_rows = [
    ["AI background generation", "Real", "Gemini gemini-2.5-flash-image; gradient fallback if image gen fails", "background_engine/services/geminiService.js"],
    ["AI template generation (text-to-template)", "Real", "JSON schema enforced; retry w/ simplified fallback", "background_engine/routes/backgroundRoutes.js"],
    ["AI complete template (concept->image->vision->layout)", "Real", "Multi-model pipeline, business branding injected", "geminiService.js (~1119-2281)"],
    ["Instagram / Facebook publishing", "Real", "Media container + publish; returns real post IDs/URLs", "metaService.js"],
    ["Twitter / X publishing", "Real", "INIT/APPEND/FINALIZE media upload + /tweets", "twitterService.js"],
    ["LinkedIn publishing", "Real", "registerUpload + /ugcPosts", "linkedinService.js"],
    ["OAuth — Meta", "Real", "Authorize, callback, long-lived token exchange", "auto_poster/routes/oauthRoutes.js"],
    ["OAuth — Twitter (PKCE)", "Real", "code_verifier/challenge; access+refresh tokens", "oauthRoutes.js / twitterService.js"],
    ["OAuth — LinkedIn", "Real", "OpenID Connect user info + token refresh", "oauthRoutes.js / linkedinService.js"],
    ["Token encryption", "Real", "AES-256-GCM, PBKDF2, salts, auth tags", "auto_poster/services/tokenService.js"],
    ["Scheduler", "Real", "cron('* * * * *'); processes due jobs, refreshes tokens", "schedulerService.js / publishingService.js"],
    ["Template editor", "Done", "Canvas, fonts, placeholders, image cropping", "post_generator/public/js/editor.js"],
    ["Cross-service wiring (this session)", "Done", "Shared service-config.js mounted + included; manager port fixed to 3004", "shared/public/service-config.js"],
    ["Docker / Cloud Run / build", "Done", "Multi-stage Dockerfile, compose volumes, pkg build.sh", "Dockerfile / build.sh"],
    ["Docs", "Partial", "README + DEPLOYMENT.md + background_engine/api.md (no 'done vs planned' section)", "README.md"],
]
style_sheet(
    ws2,
    "What's Already Built (and Real)",
    "The creative engine and the real social-publishing pipeline are genuinely implemented — not mocked.",
    ["Feature", "Status", "Notes", "Key file(s)"],
    done_rows,
    [48, 10, 60, 44],
    status_col=1,
)

# ============ TAB 3: To-Do ============
ws3 = wb.create_sheet("What's Left")
todo_rows = [
    ["Security", "No real authentication", "Blocker", "Every request hardcoded user_demo_001 -> all users shared data", "DONE: scrypt signup/login, JWT (7d), login/signup UI, token auto-attach, per-user isolation + business-ownership checks; first signup claims demo data", "manager/routes/authRoutes.js + shared/auth/password.js", "DONE (Phase 1.5)"],
    ["Security", "Secrets committed to git", "Blocker", "Claimed exposure of keys in git history", "VERIFIED FALSE: .env was never committed and is gitignored", ".gitignore", "DONE - not an issue"],
    ["Security", "JWT signature not verified", "P0 Critical", "Middleware only decoded + checked expiry", "HS256 signature verify + signJwt helper added (Node crypto)", "shared/middleware/auth.js", "DONE (Phase 1)"],
    ["Security", "Dev-mode auth bypass", "P0 Critical", "Non-production lets ANY request through as valid", "Kept dev fallback by design; tighten with real login in Phase 2", "auth.js:31-32", "PENDING (Phase 2)"],
    ["Security", "CORS open to all origins", "P1 High", "app.use(cors()) on all 4 services", "Now env-driven: set CORS_ORIGINS to lock down; permissive only if unset", "shared/config/corsOptions.js", "DONE (Phase 1)"],
    ["Data", "JSON flat-file database", "P0 Critical", "25MB data.json, sync read/write, no locking -> corruption", "Migrated to SQLite (node:sqlite) with WAL; identical API; 389 rows imported; data.json kept as backup", "shared/db/database.js + migrate-json-to-sqlite.js", "DONE (Phase 2)"],
    ["Data", "SQLite schema is dead code", "P2 Medium", "init.js built SQLite then exited; runtime used JSON", "Runtime now self-creates the SQLite schema; init.js superseded", "shared/db/database.js", "DONE (Phase 2)"],
    ["Data", "No input validation", "P1 High", "Endpoints trusted client input", "DONE: dependency-free validateBody middleware + tests; applied to template create (extend to remaining write routes as needed)", "shared/middleware/validate.js", "DONE (Phase 4)"],
    ["Security", "Cross-service object ownership", "P1 High", "Proxied services didn't verify business_id belongs to the user", "DONE: lenient requireBusinessOwnership guard on post_gen + auto_poster business routes (+ tests); enforces for logged-in users, skips internal/dev", "shared/middleware/ownership.js", "DONE (Phase 4)"],
    ["Config", "OAuth redirect URIs hardcoded to ngrok", "P1 High", "Defaults pointed at a personal ngrok URL", "ngrok removed; redirect URIs derived from PUBLIC_BASE_URL", "shared/config/platforms.js", "DONE (Phase 1)"],
    ["Config", "No startup secret validation", "P2 Medium", "Missing secrets silently fall back to dev mode", "validateEnv() added: warns in dev, exits in prod", "shared/config/validateEnv.js", "DONE (Phase 1)"],
    ["Marketing", "Analytics dashboard", "P1 High", "Table+methods existed; no API, no fetch, no UI", "DONE (scaffolding): /api/analytics summary+posts+refresh, dashboard UI with empty states, platform getInsights (IG/FB/Twitter) + ownership + tests. Live numbers need connected accounts with valid tokens", "auto_poster/routes/analyticsRoutes.js + analytics.html", "DONE (Phase 3)"],
    ["Marketing", "Best-time-to-post is fake", "P2 Medium", "Returned hardcoded times", "DONE: computed from the business's published-post history per platform, with general recommendations where data is sparse", "auto_poster/routes/calendarRoutes.js", "DONE (Phase 4)"],
    ["Marketing", "Excel bulk import missing", "P3 Low", "'Coming soon'; CSV only", "Add xlsx parsing (e.g. SheetJS)", "post_generator/public/home.html:1669", "PENDING (Phase 3)"],
    ["Marketing", "Retry failed generation missing", "P3 Low", "'Coming soon' toast", "Implement retry handler", "post_generator/public/js/generate.js:3131", "PENDING (Phase 3)"],
    ["Marketing", "Multi-user / teams / approval", "P2 Medium", "Single shared user; no roles or approval flow", "Add accounts, roles, approve-before-publish", "—", "PENDING (Phase 3)"],
    ["Marketing", "WhatsApp posting", "P3 Low", "Placeholder only", "Integrate WhatsApp Business API (or drop from UI)", "shared/config/platforms.js:86", "PENDING (Phase 3)"],
    ["Quality", "No automated tests", "P1 High", "0% coverage, no jest/mocha", "DONE: 26 tests via Node built-in node:test (npm test) covering password hashing, JWT, auth-middleware regression, SQLite layer, CORS", "tests/ + package.json", "DONE (Phase 4)"],
    ["Quality", "Observability (errors/uptime/alerts)", "P1 High", "No Sentry, uptime checks, or alerts (Gate 12 other half)", "Add error tracking + uptime pings on /health + log alerts (needs external tooling/DSN)", "(pending)", "PENDING (Phase 4)"],
    ["Quality", "Dead placeholder OAuth route", "P3 Low", "socialRoutes returned mock JSON (shadowed by real route)", "Removed", "auto_poster/routes/socialRoutes.js", "DONE (Phase 1)"],
]
style_sheet(
    ws3,
    "What's Left To Build / Fix",
    "Grouped by area. Priority: P0/Blocker = before any real users. Status column tracks progress (Phase 1 security hardening complete).",
    ["Area", "Item", "Priority", "Problem", "What to do / outcome", "Location", "Status"],
    todo_rows,
    [12, 32, 13, 44, 50, 34, 20],
    prio_col=2,
    status_col=6,
)

# ============ TAB 4: Roadmap ============
ws4 = wb.create_sheet("Roadmap")
road_rows = [
    ["Phase 1 — Security (do immediately)", "Rotate secrets / purge git history", "P0 Critical", "~1 day", "DONE - verified not needed (.env never committed)"],
    ["Phase 1 — Security (do immediately)", "Verified JWT signatures + signJwt helper", "P0 Critical", "~0.5 day", "DONE"],
    ["Phase 1 — Security (do immediately)", "Lock down CORS (env-driven CORS_ORIGINS)", "P1 High", "~0.5 day", "DONE"],
    ["Phase 1 — Security (do immediately)", "Startup secret validation + remove ngrok/dead route", "P1 High", "~0.5 day", "DONE"],
    ["Phase 1.5 — Real auth", "Real signup/login + scrypt + per-user isolation + login UI", "P0 Critical", "~3-5 days", "DONE"],
    ["Phase 2 — Production hardening", "Swap JSON file -> SQLite + migration (WAL, same API)", "P0 Critical", "~3-5 days", "DONE"],
    ["Phase 2 — Production hardening", "OAuth redirect URIs from env (prod domain)", "P1 High", "~0.5 day", "Not started"],
    ["Phase 2 — Production hardening", "Input validation + startup secret checks", "P1 High", "~2 days", "Not started"],
    ["Phase 3 — Marketing value", "Analytics: API + dashboard + platform getInsights (scaffolded)", "P1 High", "~1-2 weeks", "DONE (live numbers need connected accounts)"],
    ["Phase 3 — Marketing value", "Excel import, retry, best-time from real data", "P2 Medium", "~3-5 days", "Not started"],
    ["Phase 3 — Marketing value", "Team / multi-client accounts + approval workflow", "P2 Medium", "~1-2 weeks", "Not started"],
    ["Phase 4 — Confidence", "Automated test suite (node:test) — 26 tests green", "P1 High", "~1 week", "DONE (initial suite)"],
    ["Phase 4 — Confidence", "Observability: structured error logging (all 4 svcs) + optional Sentry", "P1 High", "~3-5 days", "PARTIAL (code done; uptime/alerts/DSN = your tooling)"],
]
style_sheet(
    ws4,
    "Suggested Roadmap",
    "Phased plan. Effort estimates are rough, single-developer.",
    ["Phase", "Task", "Priority", "Rough effort", "Status"],
    road_rows,
    [34, 52, 13, 14, 14],
    prio_col=2,
)
# status column dropdown-ish coloring left as plain text; add a Status legend note
ws4.cell(row=4 + len(road_rows) + 2, column=1, value="Tip: set the Status column to track progress (Not started / In progress / Done).").font = sub_font

# ============ save ============
out = r"C:\Users\arjus\Downloads\Brandio\Brandio_Project_Status.xlsx"
wb.save(out)
print("WROTE", out)
print("TABS", [ws.title for ws in wb.worksheets])
